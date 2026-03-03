"""Power BI-ready Excel export with fact tables, dimension tables, and metadata."""

import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.calculations import calculate_baselines, aggregate_by_period
from utils.forecasting import load_forecast_log
from utils.constants import ROAS_TARGETS, load_settings

# ---------------------------------------------------------------------------
# Theme constants (Objectif Lune)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="2D3E50", end_color="2D3E50", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
COL_WIDTH_MIN = 8
COL_WIDTH_MAX = 22


# ---------------------------------------------------------------------------
# Helper: write a DataFrame (or list-of-lists) to a worksheet with styling
# ---------------------------------------------------------------------------
def _write_sheet(ws, headers: list[str], rows: list[list], freeze: bool = True):
    """Write *headers* + *rows* to *ws* with standard Power BI styling.

    Parameters
    ----------
    ws : openpyxl Worksheet
    headers : column names (first row)
    rows : list of lists, one inner list per data row
    freeze : if True, freeze panes at A2 and enable auto-filter
    """
    # -- header row --------------------------------------------------------
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # -- data rows ---------------------------------------------------------
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Replace NaN / None with empty cell (Power BI reads as blank)
            if value is None or (isinstance(value, float) and np.isnan(value)):
                cell.value = None
            else:
                cell.value = value
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    # -- auto-width --------------------------------------------------------
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        width = max(COL_WIDTH_MIN, min(max_len + 2, COL_WIDTH_MAX))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- freeze panes & auto-filter ----------------------------------------
    if freeze and rows:
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"


def _safe_round(value, decimals=4):
    """Round numeric values; pass through non-numeric or NaN unchanged."""
    if value is None:
        return None
    if isinstance(value, float):
        if np.isnan(value):
            return None
        return round(value, decimals)
    return value


# ---------------------------------------------------------------------------
# Helper: build an aggregated fact DataFrame from filtered data
# ---------------------------------------------------------------------------
def _build_fact(filtered: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    """Aggregate raw rows and recalculate KPIs."""
    agg = (
        filtered.groupby(group_cols)
        .agg(
            spend=("spend", "sum"),
            impressions=("impressions", "sum"),
            clicks=("clicks", "sum"),
            conversions=("conversions", "sum"),
            revenue=("revenue", "sum"),
        )
        .reset_index()
    )
    agg["cpm"] = np.where(agg["impressions"] > 0, agg["spend"] / agg["impressions"] * 1000, np.nan)
    agg["ctr"] = np.where(agg["impressions"] > 0, agg["clicks"] / agg["impressions"] * 100, np.nan)
    agg["cvr"] = np.where(agg["clicks"] > 0, agg["conversions"] / agg["clicks"] * 100, np.nan)
    agg["aov"] = np.where(agg["conversions"] > 0, agg["revenue"] / agg["conversions"], np.nan)
    agg["roas"] = np.where(agg["spend"] > 0, agg["revenue"] / agg["spend"], np.nan)
    agg["cpa"] = np.where(agg["conversions"] > 0, agg["spend"] / agg["conversions"], np.nan)
    return agg


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_fact_daily(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 1 -- fact_daily: one row per date x platform x campaign_type."""
    ws = wb.active
    ws.title = "fact_daily"

    headers = [
        "date", "platform", "campaign_type",
        "spend", "impressions", "clicks", "conversions", "revenue",
        "cpm", "ctr", "cvr", "aov", "roas", "cpa",
    ]

    if filtered.empty:
        _write_sheet(ws, headers, [])
        return

    fact = _build_fact(filtered, ["date", "platform", "campaign_type"])
    fact = fact.sort_values(["date", "platform", "campaign_type"]).reset_index(drop=True)

    rows = []
    for _, r in fact.iterrows():
        rows.append([
            r["date"].strftime("%Y-%m-%d"),
            r["platform"],
            r["campaign_type"],
            _safe_round(r["spend"]),
            _safe_round(r["impressions"]),
            _safe_round(r["clicks"]),
            _safe_round(r["conversions"]),
            _safe_round(r["revenue"]),
            _safe_round(r["cpm"]),
            _safe_round(r["ctr"]),
            _safe_round(r["cvr"]),
            _safe_round(r["aov"]),
            _safe_round(r["roas"]),
            _safe_round(r["cpa"]),
        ])

    _write_sheet(ws, headers, rows)


def _sheet_fact_weekly(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 2 -- fact_weekly: weekly aggregation fact table."""
    ws = wb.create_sheet("fact_weekly")

    headers = [
        "week_start", "week_label", "platform", "campaign_type",
        "spend", "revenue", "impressions", "clicks", "conversions",
        "roas", "cpm", "ctr", "cvr", "aov",
    ]

    if filtered.empty:
        _write_sheet(ws, headers, [])
        return

    tmp = filtered.copy()
    tmp["week_start"] = tmp["date"].dt.to_period("W").apply(lambda p: p.start_time)

    fact = _build_fact(tmp, ["week_start", "platform", "campaign_type"])
    fact = fact.sort_values(["week_start", "platform", "campaign_type"]).reset_index(drop=True)

    rows = []
    for _, r in fact.iterrows():
        ws_date = pd.Timestamp(r["week_start"])
        iso_cal = ws_date.isocalendar()
        week_label = f"W{iso_cal[1]:02d} {iso_cal[0]}"
        rows.append([
            ws_date.strftime("%Y-%m-%d"),
            week_label,
            r["platform"],
            r["campaign_type"],
            _safe_round(r["spend"]),
            _safe_round(r["revenue"]),
            _safe_round(r["impressions"]),
            _safe_round(r["clicks"]),
            _safe_round(r["conversions"]),
            _safe_round(r["roas"]),
            _safe_round(r["cpm"]),
            _safe_round(r["ctr"]),
            _safe_round(r["cvr"]),
            _safe_round(r["aov"]),
        ])

    _write_sheet(ws, headers, rows)


def _sheet_fact_monthly(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 3 -- fact_monthly: monthly aggregation fact table."""
    ws = wb.create_sheet("fact_monthly")

    headers = [
        "month", "platform", "campaign_type",
        "spend", "revenue", "impressions", "clicks", "conversions",
        "roas", "cpm", "ctr", "cvr", "aov",
    ]

    if filtered.empty:
        _write_sheet(ws, headers, [])
        return

    tmp = filtered.copy()
    tmp["month_start"] = tmp["date"].dt.to_period("M").apply(lambda p: p.start_time)

    fact = _build_fact(tmp, ["month_start", "platform", "campaign_type"])
    fact = fact.sort_values(["month_start", "platform", "campaign_type"]).reset_index(drop=True)

    rows = []
    for _, r in fact.iterrows():
        rows.append([
            pd.Timestamp(r["month_start"]).strftime("%Y-%m"),
            r["platform"],
            r["campaign_type"],
            _safe_round(r["spend"]),
            _safe_round(r["revenue"]),
            _safe_round(r["impressions"]),
            _safe_round(r["clicks"]),
            _safe_round(r["conversions"]),
            _safe_round(r["roas"]),
            _safe_round(r["cpm"]),
            _safe_round(r["ctr"]),
            _safe_round(r["cvr"]),
            _safe_round(r["aov"]),
        ])

    _write_sheet(ws, headers, rows)


def _sheet_dim_platform(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 4 -- dim_platform: platform dimension with ROAS targets."""
    ws = wb.create_sheet("dim_platform")
    settings = load_settings()

    headers = ["platform", "roas_target_prospecting", "roas_target_retargeting"]

    platforms = sorted(filtered["platform"].unique()) if not filtered.empty else []
    rows = [
        [
            p,
            settings.get("roas_target_prospecting", ROAS_TARGETS.get("Prospecting", 8)),
            settings.get("roas_target_retargeting", ROAS_TARGETS.get("Retargeting", 14)),
        ]
        for p in platforms
    ]

    _write_sheet(ws, headers, rows)


def _sheet_dim_campaign_type(wb: Workbook):
    """Sheet 5 -- dim_campaign_type: campaign type dimension with ROAS targets."""
    ws = wb.create_sheet("dim_campaign_type")

    headers = ["campaign_type", "roas_target"]
    rows = [
        ["Prospecting", ROAS_TARGETS.get("Prospecting", 8)],
        ["Retargeting", ROAS_TARGETS.get("Retargeting", 14)],
    ]

    _write_sheet(ws, headers, rows)


def _sheet_dim_date(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 6 -- dim_date: calendar dimension from min to max date."""
    ws = wb.create_sheet("dim_date")

    headers = [
        "date", "year", "quarter", "month_num", "month_name",
        "week_num", "day_of_week", "day_name", "is_weekend",
    ]

    if filtered.empty:
        _write_sheet(ws, headers, [])
        return

    min_date = filtered["date"].min().normalize()
    max_date = filtered["date"].max().normalize()
    all_dates = pd.date_range(start=min_date, end=max_date, freq="D")

    rows = []
    for d in all_dates:
        iso = d.isocalendar()
        rows.append([
            d.strftime("%Y-%m-%d"),
            d.year,
            f"Q{d.quarter}",
            d.month,
            d.strftime("%B"),
            iso[1],                         # ISO week number
            iso[2],                         # Monday=1 ... Sunday=7
            d.strftime("%A"),
            iso[2] >= 6,                    # Saturday=6, Sunday=7
        ])

    _write_sheet(ws, headers, rows)


def _sheet_dim_baselines(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 7 -- dim_baselines: rolling baselines at max date."""
    ws = wb.create_sheet("dim_baselines")

    headers = [
        "platform", "campaign_type",
        "aov_60d", "cpm_14d", "ctr_14d", "cvr_14d",
        "roas_7d", "roas_14d", "roas_30d",
    ]

    if filtered.empty:
        _write_sheet(ws, headers, [])
        return

    ref_date = filtered["date"].max()
    platforms = sorted(filtered["platform"].unique())

    rows = []
    for platform in platforms:
        ctypes = sorted(
            filtered[filtered["platform"] == platform]["campaign_type"].unique()
        )
        for ctype in ctypes:
            bl = calculate_baselines(filtered, ref_date, platform=platform, campaign_type=ctype)
            rows.append([
                platform,
                ctype,
                _safe_round(bl.get("aov"), 4),
                _safe_round(bl.get("cpm"), 4),
                _safe_round(bl.get("ctr"), 4),
                _safe_round(bl.get("cvr"), 4),
                _safe_round(bl.get("roas_7d"), 4),
                _safe_round(bl.get("roas_14d"), 4),
                _safe_round(bl.get("roas_30d"), 4),
            ])

    _write_sheet(ws, headers, rows)


def _sheet_fact_forecast(wb: Workbook, filtered: pd.DataFrame):
    """Sheet 8 -- fact_forecast: forecast vs actuals (from forecast log)."""
    ws = wb.create_sheet("fact_forecast")

    headers = [
        "month", "platform", "campaign_type",
        "forecast_spend", "forecast_revenue", "forecast_roas",
        "actual_spend", "actual_revenue", "actual_roas",
        "spend_delta_pct", "revenue_delta_pct", "roas_delta_pct",
    ]

    forecast_log = load_forecast_log()
    if not forecast_log:
        _write_sheet(ws, headers, [])
        return

    rows = []
    for entry in forecast_log:
        month = entry.get("month", "")
        for proj in entry.get("projections", []):
            f_spend = proj.get("spend", 0)
            f_revenue = proj.get("revenue", 0)
            f_roas = proj.get("roas", 0)

            a_spend = None
            a_revenue = None
            a_roas = None
            spend_delta = None
            revenue_delta = None
            roas_delta = None

            # Match actuals from the data
            if not filtered.empty and month:
                try:
                    month_start = pd.Timestamp(month + "-01")
                    month_end = month_start + pd.DateOffset(months=1) - timedelta(days=1)
                    actuals = filtered[
                        (filtered["date"].between(month_start, month_end))
                        & (filtered["platform"] == proj.get("platform", ""))
                        & (filtered["campaign_type"] == proj.get("type", ""))
                    ]
                    a_spend = actuals["spend"].sum()
                    a_revenue = actuals["revenue"].sum()
                    a_roas = a_revenue / a_spend if a_spend > 0 else 0

                    if f_spend > 0:
                        spend_delta = round((a_spend - f_spend) / f_spend * 100, 4)
                    if f_revenue > 0:
                        revenue_delta = round((a_revenue - f_revenue) / f_revenue * 100, 4)
                    if f_roas > 0:
                        roas_delta = round((a_roas - f_roas) / f_roas * 100, 4)
                except Exception:
                    pass

            rows.append([
                month,
                proj.get("platform", ""),
                proj.get("type", ""),
                _safe_round(f_spend, 4),
                _safe_round(f_revenue, 4),
                _safe_round(f_roas, 4),
                _safe_round(a_spend, 4),
                _safe_round(a_revenue, 4),
                _safe_round(a_roas, 4),
                _safe_round(spend_delta, 4),
                _safe_round(revenue_delta, 4),
                _safe_round(roas_delta, 4),
            ])

    _write_sheet(ws, headers, rows)


def _sheet_meta_info(wb: Workbook, filtered: pd.DataFrame, start_date, end_date, platforms):
    """Sheet 9 -- meta_info: generation metadata and relationship map."""
    ws = wb.create_sheet("meta_info")

    # Header styling for column A (labels)
    label_font = Font(name="Calibri", bold=True, size=11, color="2D3E50")
    value_font = Font(name="Calibri", size=10)

    def _meta_row(row_num: int, label: str, value):
        cell_l = ws.cell(row=row_num, column=1, value=label)
        cell_l.font = label_font
        cell_l.border = THIN_BORDER
        cell_v = ws.cell(row=row_num, column=2, value=value)
        cell_v.font = value_font
        cell_v.border = THIN_BORDER

    # Basic metadata
    _meta_row(1, "Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    if not filtered.empty:
        d_min = filtered["date"].min().strftime("%Y-%m-%d")
        d_max = filtered["date"].max().strftime("%Y-%m-%d")
        date_range = f"{d_min} \u2014 {d_max}"
    else:
        date_range = "No data"
    _meta_row(2, "Date Range", date_range)

    plat_list = ", ".join(sorted(filtered["platform"].unique())) if not filtered.empty else ""
    _meta_row(3, "Platforms", plat_list)

    _meta_row(4, "Total Rows", len(filtered))

    # Relationship map
    row_num = 6
    header_cell = ws.cell(row=row_num, column=1, value="Relationships")
    header_cell.font = HEADER_FONT
    header_cell.fill = HEADER_FILL
    header_cell.border = THIN_BORDER
    ws.cell(row=row_num, column=2).fill = HEADER_FILL
    ws.cell(row=row_num, column=2).border = THIN_BORDER

    relationships = [
        ("fact_daily.date", "dim_date.date"),
        ("fact_daily.platform", "dim_platform.platform"),
        ("fact_daily.campaign_type", "dim_campaign_type.campaign_type"),
        ("fact_daily.platform + campaign_type", "dim_baselines.platform + campaign_type"),
    ]
    for rel_from, rel_to in relationships:
        row_num += 1
        c1 = ws.cell(row=row_num, column=1, value=rel_from)
        c1.font = value_font
        c1.border = THIN_BORDER
        c2 = ws.cell(row=row_num, column=2, value=rel_to)
        c2.font = value_font
        c2.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 42


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_powerbi_export(
    df: pd.DataFrame,
    start_date=None,
    end_date=None,
    platforms: list = None,
) -> BytesIO:
    """Generate a Power BI-ready Excel workbook with fact and dimension tables.

    Sheets
    ------
    1. fact_daily     -- granular fact table (date x platform x campaign_type)
    2. fact_weekly    -- weekly aggregation fact table
    3. fact_monthly   -- monthly aggregation fact table
    4. dim_platform   -- platform dimension with ROAS targets
    5. dim_campaign_type -- campaign type dimension with ROAS targets
    6. dim_date       -- full calendar dimension (min to max date)
    7. dim_baselines  -- rolling baseline values at latest date
    8. fact_forecast  -- forecast vs actuals from forecast log
    9. meta_info      -- generation metadata and join-key reference

    Parameters
    ----------
    df : pd.DataFrame
        Raw performance data with standard columns.
    start_date, end_date : date-like, optional
        Filter bounds.
    platforms : list[str], optional
        Restrict to these platforms.

    Returns
    -------
    BytesIO
        In-memory Excel file ready for download.
    """
    wb = Workbook()

    # -- apply filters -----------------------------------------------------
    filtered = df.copy()
    if start_date is not None:
        filtered = filtered[filtered["date"] >= pd.Timestamp(start_date)]
    if end_date is not None:
        filtered = filtered[filtered["date"] <= pd.Timestamp(end_date)]
    if platforms:
        filtered = filtered[filtered["platform"].isin(platforms)]

    # -- build each sheet --------------------------------------------------
    _sheet_fact_daily(wb, filtered)
    _sheet_fact_weekly(wb, filtered)
    _sheet_fact_monthly(wb, filtered)
    _sheet_dim_platform(wb, filtered)
    _sheet_dim_campaign_type(wb)
    _sheet_dim_date(wb, filtered)
    _sheet_dim_baselines(wb, filtered)
    _sheet_fact_forecast(wb, filtered)
    _sheet_meta_info(wb, filtered, start_date, end_date, platforms)

    # -- serialize ---------------------------------------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
