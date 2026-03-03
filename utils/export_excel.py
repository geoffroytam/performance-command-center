"""Excel report generation with professional formatting, conditional styling, and multi-sheet layout."""

import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from utils.calculations import (
    calculate_baselines,
    aggregate_by_period,
    aggregate_for_date,
)
from utils.constants import COLORS, ROAS_TARGETS, load_settings


# ---------------------------------------------------------------------------
# Theme colours (Objectif Lune palette)
# ---------------------------------------------------------------------------
_CLR_HEADER_BG = "2D3E50"   # midnight
_CLR_ACCENT = "4A6FA5"      # primary blue
_CLR_GREEN = "6B8F71"
_CLR_RED = "C45C4A"
_CLR_WHITE = "FAFAF7"       # warm white
_CLR_CREAM = "F5F0E8"
_CLR_LIGHT_GRAY = "F0EDE6"
_CLR_GRAY_TEXT = "7A7A72"
_CLR_BORDER = "D5D8DC"

# Light tint backgrounds for conditional formatting
_CLR_GREEN_LIGHT = "E8F0E9"
_CLR_RED_LIGHT = "F5E0DC"

# ---------------------------------------------------------------------------
# Reusable openpyxl style objects
# ---------------------------------------------------------------------------
_BORDER_THIN = Border(
    left=Side(style="thin", color=_CLR_BORDER),
    right=Side(style="thin", color=_CLR_BORDER),
    top=Side(style="thin", color=_CLR_BORDER),
    bottom=Side(style="thin", color=_CLR_BORDER),
)

_FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_FONT_DATA = Font(name="Calibri", size=10, color="333333")
_FONT_DATA_BOLD = Font(name="Calibri", size=10, bold=True, color="333333")
_FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=_CLR_HEADER_BG)
_FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=_CLR_HEADER_BG)
_FONT_KPI_LABEL = Font(name="Calibri", size=9, color=_CLR_GRAY_TEXT)
_FONT_KPI_VALUE = Font(name="Calibri", size=14, bold=True, color=_CLR_HEADER_BG)
_FONT_FOOTER = Font(name="Calibri", size=8, italic=True, color=_CLR_GRAY_TEXT)
_FONT_WOW_GREEN = Font(name="Calibri", size=10, color=_CLR_GREEN)
_FONT_WOW_RED = Font(name="Calibri", size=10, color=_CLR_RED)

_FILL_HEADER = PatternFill(start_color=_CLR_HEADER_BG, end_color=_CLR_HEADER_BG, fill_type="solid")
_FILL_WHITE = PatternFill(start_color=_CLR_WHITE, end_color=_CLR_WHITE, fill_type="solid")
_FILL_CREAM = PatternFill(start_color=_CLR_CREAM, end_color=_CLR_CREAM, fill_type="solid")
_FILL_GREEN_LIGHT = PatternFill(start_color=_CLR_GREEN_LIGHT, end_color=_CLR_GREEN_LIGHT, fill_type="solid")
_FILL_RED_LIGHT = PatternFill(start_color=_CLR_RED_LIGHT, end_color=_CLR_RED_LIGHT, fill_type="solid")
_FILL_KPI_BG = PatternFill(start_color=_CLR_LIGHT_GRAY, end_color=_CLR_LIGHT_GRAY, fill_type="solid")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)

# ---------------------------------------------------------------------------
# Number format strings (Excel-native, not Python formatting)
# ---------------------------------------------------------------------------
_FMT_CURRENCY = 'R$ #,##0'
_FMT_CURRENCY_DEC = 'R$ #,##0.00'
_FMT_ROAS = '0.0'
_FMT_PCT = '0.00%'
_FMT_NUMBER = '#,##0'
_FMT_WOW_PCT = '+0.0%;-0.0%'
_FMT_DATE = 'DD/MM/YYYY'


# ═══════════════════════════════════════════════════════════════════════════
# Helper functions
# ═══════════════════════════════════════════════════════════════════════════

def _get_roas_target(campaign_type: str) -> float:
    """Return the ROAS target for a given campaign type."""
    settings = load_settings()
    if campaign_type and "retarg" in str(campaign_type).lower():
        return settings.get("roas_target_retargeting", 14)
    return settings.get("roas_target_prospecting", 8)


def _apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Apply multiple style attributes to a cell in one call."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _set_col_widths(ws, col_widths: dict):
    """Set column widths from a dict of {column_letter_or_index: width}."""
    for col, width in col_widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = max(8, min(width, 22))


def _auto_col_widths(ws, min_width: int = 10, max_width: int = 22):
    """Auto-adjust column widths based on content, with min/max bounds."""
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                lengths.append(len(val_str))
        best = max(lengths) + 3 if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(best, max_width))


def _write_header_row(ws, row: int, columns: list, col_widths: dict = None):
    """Write a styled header row and optionally set column widths."""
    for c_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=c_idx, value=col_name)
        _apply_cell_style(
            cell,
            font=_FONT_HEADER,
            fill=_FILL_HEADER,
            alignment=_ALIGN_CENTER,
            border=_BORDER_THIN,
        )
    if col_widths:
        _set_col_widths(ws, col_widths)


def _write_data_rows(
    ws,
    data_rows: list[list],
    start_row: int,
    col_formats: list = None,
    alignments: list = None,
):
    """
    Write data rows with alternating shading, borders, and number formats.

    col_formats:  list of Excel number format strings (or None) per column.
    alignments:   list of Alignment objects (or None) per column.
    """
    num_cols = len(data_rows[0]) if data_rows else 0

    for r_offset, row_data in enumerate(data_rows):
        row_num = start_row + r_offset
        fill = _FILL_CREAM if r_offset % 2 == 0 else _FILL_WHITE
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_num, column=c_idx + 1)
            # Handle NaN / None
            if isinstance(value, float) and (pd.isna(value) or np.isnan(value)):
                cell.value = None
            else:
                cell.value = value

            nf = col_formats[c_idx] if col_formats and c_idx < len(col_formats) else None
            al = alignments[c_idx] if alignments and c_idx < len(alignments) else _ALIGN_CENTER

            _apply_cell_style(
                cell,
                font=_FONT_DATA,
                fill=fill,
                alignment=al,
                border=_BORDER_THIN,
                number_format=nf,
            )


def _apply_roas_conditional(ws, roas_col: int, type_col: int, start_row: int, end_row: int):
    """Colour-code ROAS cells: green if >= target, red if below."""
    for row in range(start_row, end_row + 1):
        roas_cell = ws.cell(row=row, column=roas_col)
        type_cell = ws.cell(row=row, column=type_col)
        try:
            roas_val = float(roas_cell.value) if roas_cell.value is not None else None
        except (ValueError, TypeError):
            continue
        if roas_val is None:
            continue

        target = _get_roas_target(type_cell.value)
        if roas_val >= target:
            roas_cell.fill = _FILL_GREEN_LIGHT
        else:
            roas_cell.fill = _FILL_RED_LIGHT


def _apply_wow_formatting(ws, col: int, start_row: int, end_row: int, higher_is_better: bool = True):
    """Apply green/red font to WoW or MoM percentage-change columns."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)
        try:
            val = float(cell.value) if cell.value is not None else None
        except (ValueError, TypeError):
            continue
        if val is None:
            continue

        is_good = (val >= 0) if higher_is_better else (val <= 0)
        cell.font = _FONT_WOW_GREEN if is_good else _FONT_WOW_RED


def _add_autofilter(ws, header_row: int, num_cols: int, num_data_rows: int):
    """Set auto-filter on the header row spanning all data."""
    last_col = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + num_data_rows}"


# ═══════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════

def _build_executive_summary(wb, filtered: pd.DataFrame, start_date, end_date, full_df: pd.DataFrame):
    """Sheet 1 -- Executive Summary with KPI cards and platform breakdown."""
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = _CLR_ACCENT

    date_str = f"{pd.Timestamp(start_date).strftime('%d/%m/%Y')}  —  {pd.Timestamp(end_date).strftime('%d/%m/%Y')}"

    # ── Title ─────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value=f"Performance Report  —  {date_str}")
    _apply_cell_style(title_cell, font=_FONT_TITLE, alignment=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 36

    # ── KPI Cards ─────────────────────────────────────────────
    total_spend = filtered["spend"].sum()
    total_revenue = filtered["revenue"].sum()
    total_orders = filtered["conversions"].sum()
    overall_roas = total_revenue / total_spend if total_spend > 0 else 0
    overall_aov = total_revenue / total_orders if total_orders > 0 else 0

    kpi_labels = ["Total Spend", "Total Revenue", "ROAS", "Orders", "AOV"]
    kpi_values = [total_spend, total_revenue, overall_roas, int(total_orders), overall_aov]
    kpi_formats = [_FMT_CURRENCY, _FMT_CURRENCY, _FMT_ROAS, _FMT_NUMBER, _FMT_CURRENCY_DEC]

    ws.cell(row=3, column=1, value="KEY METRICS").font = _FONT_SECTION
    label_row = 5
    value_row = 6

    for i, (label, value, fmt) in enumerate(zip(kpi_labels, kpi_values, kpi_formats)):
        col = i * 2 + 1  # columns 1, 3, 5, 7, 9

        # Label cell
        lbl_cell = ws.cell(row=label_row, column=col, value=label)
        _apply_cell_style(lbl_cell, font=_FONT_KPI_LABEL, alignment=_ALIGN_CENTER)
        # Merge two columns for width
        ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=col + 1)
        ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row, end_column=col + 1)

        # Value cell
        val_cell = ws.cell(row=value_row, column=col, value=value)
        _apply_cell_style(
            val_cell,
            font=_FONT_KPI_VALUE,
            fill=_FILL_KPI_BG,
            alignment=_ALIGN_CENTER,
            border=_BORDER_THIN,
            number_format=fmt,
        )
        # Also style the merged companion cell border
        companion = ws.cell(row=value_row, column=col + 1)
        _apply_cell_style(companion, fill=_FILL_KPI_BG, border=_BORDER_THIN)

    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 30

    # ── Platform Breakdown Table ──────────────────────────────
    section_row = 9
    ws.cell(row=section_row, column=1, value="PLATFORM BREAKDOWN").font = _FONT_SECTION

    header_row = 11
    columns = ["Platform", "Type", "Spend", "Revenue", "ROAS", "Orders", "CPM", "CTR", "CVR"]
    _write_header_row(ws, header_row, columns)

    # Aggregate by platform + campaign_type
    agg = (
        filtered.groupby(["platform", "campaign_type"])
        .agg(
            spend=("spend", "sum"),
            revenue=("revenue", "sum"),
            impressions=("impressions", "sum"),
            clicks=("clicks", "sum"),
            conversions=("conversions", "sum"),
        )
        .reset_index()
        .sort_values(["platform", "campaign_type"])
    )
    agg["roas"] = np.where(agg["spend"] > 0, agg["revenue"] / agg["spend"], np.nan)
    agg["cpm"] = np.where(agg["impressions"] > 0, agg["spend"] / agg["impressions"] * 1000, np.nan)
    agg["ctr"] = np.where(agg["impressions"] > 0, agg["clicks"] / agg["impressions"], np.nan)
    agg["cvr"] = np.where(agg["clicks"] > 0, agg["conversions"] / agg["clicks"], np.nan)

    data_rows = []
    for _, r in agg.iterrows():
        data_rows.append([
            r["platform"],
            r["campaign_type"],
            r["spend"],
            r["revenue"],
            r["roas"],
            int(r["conversions"]),
            r["cpm"],
            r["ctr"],
            r["cvr"],
        ])

    col_formats = [
        None,             # Platform
        None,             # Type
        _FMT_CURRENCY,    # Spend
        _FMT_CURRENCY,    # Revenue
        _FMT_ROAS,        # ROAS
        _FMT_NUMBER,      # Orders
        _FMT_CURRENCY_DEC,# CPM
        _FMT_PCT,         # CTR
        _FMT_PCT,         # CVR
    ]
    col_alignments = [
        _ALIGN_LEFT,      # Platform
        _ALIGN_LEFT,      # Type
        _ALIGN_RIGHT,     # Spend
        _ALIGN_RIGHT,     # Revenue
        _ALIGN_CENTER,    # ROAS
        _ALIGN_CENTER,    # Orders
        _ALIGN_RIGHT,     # CPM
        _ALIGN_CENTER,    # CTR
        _ALIGN_CENTER,    # CVR
    ]

    data_start = header_row + 1
    _write_data_rows(ws, data_rows, data_start, col_formats, col_alignments)

    # ROAS conditional formatting
    roas_col = 5  # column E
    type_col = 2  # column B
    _apply_roas_conditional(ws, roas_col, type_col, data_start, data_start + len(data_rows) - 1)

    # ── Totals row ────────────────────────────────────────────
    totals_row = data_start + len(data_rows)
    total_imp = agg["impressions"].sum()
    total_clk = agg["clicks"].sum()
    totals = [
        "TOTAL", "",
        total_spend, total_revenue, overall_roas, int(total_orders),
        (total_spend / total_imp * 1000) if total_imp > 0 else None,
        (total_clk / total_imp) if total_imp > 0 else None,
        (total_orders / total_clk) if total_clk > 0 else None,
    ]
    for c_idx, val in enumerate(totals):
        cell = ws.cell(row=totals_row, column=c_idx + 1, value=val)
        nf = col_formats[c_idx]
        _apply_cell_style(
            cell,
            font=_FONT_DATA_BOLD,
            fill=_FILL_KPI_BG,
            alignment=col_alignments[c_idx],
            border=_BORDER_THIN,
            number_format=nf,
        )

    # ── Footer ────────────────────────────────────────────────
    footer_row = totals_row + 2
    ws.cell(row=footer_row, column=1, value=f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = _FONT_FOOTER

    # Column widths
    widths = {1: 14, 2: 16, 3: 16, 4: 16, 5: 10, 6: 10, 7: 14, 8: 10, 9: 10}
    _set_col_widths(ws, widths)


def _build_daily_data(wb, filtered: pd.DataFrame):
    """Sheet 2 -- Daily Data with full formatting."""
    ws = wb.create_sheet("Daily Data")
    ws.sheet_properties.tabColor = _CLR_ACCENT

    daily = aggregate_by_period(filtered, "daily", group_cols=["platform", "campaign_type"])
    daily = daily.sort_values(["period", "platform", "campaign_type"]).reset_index(drop=True)

    columns = ["Date", "Platform", "Type", "Spend", "Revenue", "ROAS",
               "Orders", "Impressions", "Clicks", "CPM", "CTR", "CVR", "AOV", "CPA"]

    col_formats = [
        _FMT_DATE,         # Date
        None,              # Platform
        None,              # Type
        _FMT_CURRENCY,     # Spend
        _FMT_CURRENCY,     # Revenue
        _FMT_ROAS,         # ROAS
        _FMT_NUMBER,       # Orders
        _FMT_NUMBER,       # Impressions
        _FMT_NUMBER,       # Clicks
        _FMT_CURRENCY_DEC, # CPM
        _FMT_PCT,          # CTR
        _FMT_PCT,          # CVR
        _FMT_CURRENCY_DEC, # AOV
        _FMT_CURRENCY_DEC, # CPA
    ]
    col_alignments = [
        _ALIGN_CENTER,  # Date
        _ALIGN_LEFT,    # Platform
        _ALIGN_LEFT,    # Type
        _ALIGN_RIGHT,   # Spend
        _ALIGN_RIGHT,   # Revenue
        _ALIGN_CENTER,  # ROAS
        _ALIGN_CENTER,  # Orders
        _ALIGN_RIGHT,   # Impressions
        _ALIGN_RIGHT,   # Clicks
        _ALIGN_RIGHT,   # CPM
        _ALIGN_CENTER,  # CTR
        _ALIGN_CENTER,  # CVR
        _ALIGN_RIGHT,   # AOV
        _ALIGN_RIGHT,   # CPA
    ]

    _write_header_row(ws, 1, columns)

    data_rows = []
    for _, r in daily.iterrows():
        # CTR and CVR: convert from percentage (e.g. 2.5) to decimal (0.025) for Excel % format
        ctr_dec = r.get("ctr", np.nan)
        cvr_dec = r.get("cvr", np.nan)
        ctr_dec = ctr_dec / 100 if pd.notna(ctr_dec) else np.nan
        cvr_dec = cvr_dec / 100 if pd.notna(cvr_dec) else np.nan

        data_rows.append([
            r["period"],  # Date (datetime, formatted by Excel)
            r.get("platform", ""),
            r.get("campaign_type", ""),
            r.get("spend", np.nan),
            r.get("revenue", np.nan),
            r.get("roas", np.nan),
            int(r.get("conversions", 0)) if pd.notna(r.get("conversions")) else None,
            int(r.get("impressions", 0)) if pd.notna(r.get("impressions")) else None,
            int(r.get("clicks", 0)) if pd.notna(r.get("clicks")) else None,
            r.get("cpm", np.nan),
            ctr_dec,
            cvr_dec,
            r.get("aov", np.nan),
            r.get("cpa", np.nan),
        ])

    _write_data_rows(ws, data_rows, 2, col_formats, col_alignments)

    # ROAS conditional formatting (col 6 = ROAS, col 3 = Type)
    if data_rows:
        _apply_roas_conditional(ws, 6, 3, 2, 1 + len(data_rows))

    # Freeze panes, auto-filter
    ws.freeze_panes = "A2"
    _add_autofilter(ws, 1, len(columns), len(data_rows))

    # Column widths
    widths = {1: 12, 2: 12, 3: 14, 4: 14, 5: 14, 6: 10, 7: 10, 8: 14, 9: 10, 10: 12, 11: 10, 12: 10, 13: 12, 14: 12}
    _set_col_widths(ws, widths)


def _build_weekly_trends(wb, filtered: pd.DataFrame):
    """Sheet 3 -- Weekly Trends with WoW deltas."""
    ws = wb.create_sheet("Weekly Trends")
    ws.sheet_properties.tabColor = _CLR_ACCENT

    weekly = aggregate_by_period(filtered, "weekly", group_cols=["platform", "campaign_type"])
    weekly = weekly.sort_values(["period", "platform", "campaign_type"]).reset_index(drop=True)

    # Compute WoW deltas as fractional change (for Excel % format)
    for kpi in ["roas", "cpm", "ctr", "cvr", "aov"]:
        weekly[f"{kpi}_wow"] = weekly.groupby(["platform", "campaign_type"])[kpi].pct_change()

    columns = ["Week", "Platform", "Type", "Spend", "Revenue", "ROAS", "ROAS WoW",
               "Orders", "CPM", "CPM WoW", "CTR", "CVR", "AOV"]

    col_formats = [
        None,              # Week (text)
        None,              # Platform
        None,              # Type
        _FMT_CURRENCY,     # Spend
        _FMT_CURRENCY,     # Revenue
        _FMT_ROAS,         # ROAS
        _FMT_WOW_PCT,      # ROAS WoW
        _FMT_NUMBER,       # Orders
        _FMT_CURRENCY_DEC, # CPM
        _FMT_WOW_PCT,      # CPM WoW
        _FMT_PCT,          # CTR
        _FMT_PCT,          # CVR
        _FMT_CURRENCY_DEC, # AOV
    ]
    col_alignments = [
        _ALIGN_CENTER,  # Week
        _ALIGN_LEFT,    # Platform
        _ALIGN_LEFT,    # Type
        _ALIGN_RIGHT,   # Spend
        _ALIGN_RIGHT,   # Revenue
        _ALIGN_CENTER,  # ROAS
        _ALIGN_CENTER,  # ROAS WoW
        _ALIGN_CENTER,  # Orders
        _ALIGN_RIGHT,   # CPM
        _ALIGN_CENTER,  # CPM WoW
        _ALIGN_CENTER,  # CTR
        _ALIGN_CENTER,  # CVR
        _ALIGN_RIGHT,   # AOV
    ]

    _write_header_row(ws, 1, columns)

    data_rows = []
    for _, r in weekly.iterrows():
        week_label = r["period"].strftime("W%V %Y") if pd.notna(r["period"]) else ""
        ctr_dec = r.get("ctr", np.nan)
        cvr_dec = r.get("cvr", np.nan)
        ctr_dec = ctr_dec / 100 if pd.notna(ctr_dec) else np.nan
        cvr_dec = cvr_dec / 100 if pd.notna(cvr_dec) else np.nan

        data_rows.append([
            week_label,
            r.get("platform", ""),
            r.get("campaign_type", ""),
            r.get("spend", np.nan),
            r.get("revenue", np.nan),
            r.get("roas", np.nan),
            r.get("roas_wow", np.nan),
            int(r.get("conversions", 0)) if pd.notna(r.get("conversions")) else None,
            r.get("cpm", np.nan),
            r.get("cpm_wow", np.nan),
            ctr_dec,
            cvr_dec,
            r.get("aov", np.nan),
        ])

    _write_data_rows(ws, data_rows, 2, col_formats, col_alignments)

    # ROAS conditional formatting (col 6 = ROAS, col 3 = Type)
    if data_rows:
        end_row = 1 + len(data_rows)
        _apply_roas_conditional(ws, 6, 3, 2, end_row)
        # WoW formatting: ROAS WoW (col 7) - higher is better
        _apply_wow_formatting(ws, 7, 2, end_row, higher_is_better=True)
        # CPM WoW (col 10) - lower is better
        _apply_wow_formatting(ws, 10, 2, end_row, higher_is_better=False)

    ws.freeze_panes = "A2"
    _add_autofilter(ws, 1, len(columns), len(data_rows))

    widths = {1: 12, 2: 12, 3: 14, 4: 14, 5: 14, 6: 10, 7: 12, 8: 10, 9: 12, 10: 12, 11: 10, 12: 10, 13: 12}
    _set_col_widths(ws, widths)


def _build_monthly_summary(wb, filtered: pd.DataFrame):
    """Sheet 4 -- Monthly Summary with MoM change."""
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = _CLR_ACCENT

    monthly = aggregate_by_period(filtered, "monthly", group_cols=["platform", "campaign_type"])
    monthly = monthly.sort_values(["period", "platform", "campaign_type"]).reset_index(drop=True)

    # MoM delta (fractional)
    monthly["roas_mom"] = monthly.groupby(["platform", "campaign_type"])["roas"].pct_change()

    columns = ["Month", "Platform", "Type", "Spend", "Revenue", "ROAS", "ROAS MoM",
               "Orders", "CPM", "CTR", "CVR", "AOV"]

    col_formats = [
        None,              # Month (text)
        None,              # Platform
        None,              # Type
        _FMT_CURRENCY,     # Spend
        _FMT_CURRENCY,     # Revenue
        _FMT_ROAS,         # ROAS
        _FMT_WOW_PCT,      # ROAS MoM
        _FMT_NUMBER,       # Orders
        _FMT_CURRENCY_DEC, # CPM
        _FMT_PCT,          # CTR
        _FMT_PCT,          # CVR
        _FMT_CURRENCY_DEC, # AOV
    ]
    col_alignments = [
        _ALIGN_CENTER,  # Month
        _ALIGN_LEFT,    # Platform
        _ALIGN_LEFT,    # Type
        _ALIGN_RIGHT,   # Spend
        _ALIGN_RIGHT,   # Revenue
        _ALIGN_CENTER,  # ROAS
        _ALIGN_CENTER,  # ROAS MoM
        _ALIGN_CENTER,  # Orders
        _ALIGN_RIGHT,   # CPM
        _ALIGN_CENTER,  # CTR
        _ALIGN_CENTER,  # CVR
        _ALIGN_RIGHT,   # AOV
    ]

    _write_header_row(ws, 1, columns)

    data_rows = []
    for _, r in monthly.iterrows():
        month_label = r["period"].strftime("%b %Y") if pd.notna(r["period"]) else ""
        ctr_dec = r.get("ctr", np.nan)
        cvr_dec = r.get("cvr", np.nan)
        ctr_dec = ctr_dec / 100 if pd.notna(ctr_dec) else np.nan
        cvr_dec = cvr_dec / 100 if pd.notna(cvr_dec) else np.nan

        data_rows.append([
            month_label,
            r.get("platform", ""),
            r.get("campaign_type", ""),
            r.get("spend", np.nan),
            r.get("revenue", np.nan),
            r.get("roas", np.nan),
            r.get("roas_mom", np.nan),
            int(r.get("conversions", 0)) if pd.notna(r.get("conversions")) else None,
            r.get("cpm", np.nan),
            ctr_dec,
            cvr_dec,
            r.get("aov", np.nan),
        ])

    _write_data_rows(ws, data_rows, 2, col_formats, col_alignments)

    if data_rows:
        end_row = 1 + len(data_rows)
        _apply_roas_conditional(ws, 6, 3, 2, end_row)
        _apply_wow_formatting(ws, 7, 2, end_row, higher_is_better=True)

    ws.freeze_panes = "A2"
    _add_autofilter(ws, 1, len(columns), len(data_rows))

    widths = {1: 12, 2: 12, 3: 14, 4: 14, 5: 14, 6: 10, 7: 12, 8: 10, 9: 12, 10: 10, 11: 10, 12: 12}
    _set_col_widths(ws, widths)


def _build_baselines(wb, filtered: pd.DataFrame, end_date):
    """Sheet 5 -- Current baseline values per platform/type."""
    ws = wb.create_sheet("Baselines")
    ws.sheet_properties.tabColor = _CLR_GREEN

    ref_date = pd.Timestamp(end_date)

    columns = ["Platform", "Type", "AOV (60d)", "CPM (14d)", "CTR (14d)",
               "CVR (14d)", "ROAS (7d)", "ROAS (14d)", "ROAS (30d)"]

    col_formats = [
        None,              # Platform
        None,              # Type
        _FMT_CURRENCY_DEC, # AOV
        _FMT_CURRENCY_DEC, # CPM
        _FMT_PCT,          # CTR
        _FMT_PCT,          # CVR
        _FMT_ROAS,         # ROAS 7d
        _FMT_ROAS,         # ROAS 14d
        _FMT_ROAS,         # ROAS 30d
    ]
    col_alignments = [
        _ALIGN_LEFT,    # Platform
        _ALIGN_LEFT,    # Type
        _ALIGN_RIGHT,   # AOV
        _ALIGN_RIGHT,   # CPM
        _ALIGN_CENTER,  # CTR
        _ALIGN_CENTER,  # CVR
        _ALIGN_CENTER,  # ROAS 7d
        _ALIGN_CENTER,  # ROAS 14d
        _ALIGN_CENTER,  # ROAS 30d
    ]

    _write_header_row(ws, 1, columns)

    data_rows = []
    for platform in sorted(filtered["platform"].unique()):
        for ctype in sorted(filtered[filtered["platform"] == platform]["campaign_type"].unique()):
            b = calculate_baselines(filtered, ref_date, platform=platform, campaign_type=ctype)

            # CTR and CVR baselines come as percentages (e.g. 2.5) -- convert to decimal for Excel
            ctr_val = b.get("ctr", np.nan)
            cvr_val = b.get("cvr", np.nan)
            ctr_val = ctr_val / 100 if pd.notna(ctr_val) else np.nan
            cvr_val = cvr_val / 100 if pd.notna(cvr_val) else np.nan

            data_rows.append([
                platform,
                ctype,
                b.get("aov", np.nan),
                b.get("cpm", np.nan),
                ctr_val,
                cvr_val,
                b.get("roas_7d", np.nan),
                b.get("roas_14d", np.nan),
                b.get("roas_30d", np.nan),
            ])

    if data_rows:
        _write_data_rows(ws, data_rows, 2, col_formats, col_alignments)

    ws.freeze_panes = "A2"
    if data_rows:
        _add_autofilter(ws, 1, len(columns), len(data_rows))

    widths = {1: 14, 2: 16, 3: 14, 4: 14, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12}
    _set_col_widths(ws, widths)


# ═══════════════════════════════════════════════════════════════════════════
# Main entry point
# ═══════════════════════════════════════════════════════════════════════════

def generate_excel_report(
    df: pd.DataFrame,
    start_date,
    end_date,
    platforms: list = None,
    include_charts: bool = True,
) -> BytesIO:
    """
    Generate a professionally formatted Excel report with five sheets:
      1. Executive Summary  -- KPI cards + platform breakdown
      2. Daily Data         -- full daily granularity
      3. Weekly Trends      -- weekly aggregation with WoW deltas
      4. Monthly Summary    -- monthly aggregation with MoM deltas
      5. Baselines          -- rolling baseline values

    Returns a BytesIO buffer ready for Streamlit download.
    """
    wb = Workbook()

    # ── Filter data to the requested window ───────────────────
    mask = df["date"].between(pd.Timestamp(start_date), pd.Timestamp(end_date))
    if platforms:
        mask = mask & df["platform"].isin(platforms)
    filtered = df[mask].copy()

    if filtered.empty:
        ws = wb.active
        ws.title = "No Data"
        cell = ws.cell(row=2, column=2, value="No data available for the selected filters.")
        cell.font = Font(name="Calibri", size=12, color=_CLR_GRAY_TEXT)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ── Build each sheet ──────────────────────────────────────
    _build_executive_summary(wb, filtered, start_date, end_date, df)
    _build_daily_data(wb, filtered)
    _build_weekly_trends(wb, filtered)
    _build_monthly_summary(wb, filtered)
    _build_baselines(wb, filtered, end_date)

    # ── Save to buffer ────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
